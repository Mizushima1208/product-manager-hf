"""Excel export service."""
import tempfile
from datetime import datetime
from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from core import database


def export_receipts_to_excel():
    """Export all receipts to Excel file."""
    receipts = database.get_all_receipts()
    if not receipts:
        raise HTTPException(status_code=400, detail="No receipts to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "レシート一覧"

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["ID", "店舗名", "購入日", "合計金額", "商品一覧", "ファイル名", "OCRエンジン", "登録日時"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row, receipt in enumerate(receipts, 2):
        ws.cell(row=row, column=1, value=receipt["id"]).border = thin_border
        ws.cell(row=row, column=2, value=receipt.get("store_name", "")).border = thin_border
        ws.cell(row=row, column=3, value=receipt.get("purchase_date", "")).border = thin_border
        ws.cell(row=row, column=4, value=receipt.get("total_amount", "")).border = thin_border
        ws.cell(row=row, column=5, value=", ".join(receipt.get("items", []))).border = thin_border
        ws.cell(row=row, column=6, value=receipt.get("file_name", "")).border = thin_border
        ws.cell(row=row, column=7, value=receipt.get("ocr_engine", "")).border = thin_border
        ws.cell(row=row, column=8, value=receipt.get("created_at", "")).border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
